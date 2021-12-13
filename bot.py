import os
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from telegram import Update
from telegram.ext import Updater, MessageHandler, CallbackContext, Filters, CommandHandler, ConversationHandler
import datetime
import openpyxl
import database
import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

TYPING_AMOUNT, TYPING_TYPE, TYPING_TIME, TYPING_COMMENT = range(4)
GET_DAY_LOOKUP = GET_CATEGORY = GET_MONTH_LOOKUP = 0

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)


def parse_date(date):
    now = datetime.datetime.now()
    if len(date) <= 2:
        real_date = str(now.year) + '-' + str(now.month) + '-' + date
        return real_date
    for fmt in ("%Y-%m-%d", "%Y %m %d", "%Y.%m.%d", "%Y,%m,%d", "%Y\\%m\\%d", "%Y/%m/%d"):
        try:
            formatted = datetime.datetime.strptime(date, fmt)
            return formatted.strftime("%Y-%m-%d")
        except ValueError:
            pass


def reset_budget(username):
    if datetime.datetime.now().day == 1:
        database.set_all_users("spent", username, 0)


def get_budget_and_spent(username):
    planned = database.lookup_all_users("budget", username).fetched_info[0][0]
    spent = database.lookup_all_users("spent", username).fetched_info[0][0]

    return planned, spent


def bot_check_stop_in_lookup(line, update: Update, context: CallbackContext):
    if "urfin_end" in line.lower():
        context.bot.send_message(chat_id=update.effective_chat.id, text="Stop lookup!")
        return True
    else:
        return False


def bot_check_stop_in_add(line, update: Update, context: CallbackContext):
    if "urfin_end" in line.lower:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Stop add!")
        return True
    else:
        return False


def bot_set_budget(update: Update, context: CallbackContext):
    try:
        amount = update.message.text.split()[1]
        context.bot.send_message(chat_id=update.effective_chat.id, text="Got it, updating your budget!")
        try:
            database.set_all_users("budget", update.effective_user.username.lower().lower(), amount)
            context.bot.send_message(chat_id=update.effective_chat.id,
                                     text="Done! Your monthly budget is now {0}.".format(amount))
        except database.psycopg2.Error:
            context.bot.send_message(chat_id=update.effective_chat.id, text="Sorry, there's something wrong. :(")
    except IndexError:
        context.bot.send_message(chat_id=update.effective_chat.id,
                                 text="It seems that you forgot to type your budget, please, try again")


def bot_category_lookup(update: Update, context: CallbackContext):
    context.bot.send_message(chat_id=update.effective_chat.id, text="Please, enter the category you want to get "
                                                                    "information on. You can specify the order of "
                                                                    "information by passing column name after "
                                                                    "the month (e.g. food amount).\n\n"
                                                                    "COLUMNS AVAILABLE: amount, day, user_time")
    return GET_CATEGORY


def bot_day_lookup(update: Update, context: CallbackContext):  # return to user info from particular day
    context.bot.send_message(chat_id=update.effective_chat.id, text="Please, enter the date you want to get "
                                                                    "information on. You can enter the date in two "
                                                                    "ways: a single or two digits meaning day of "
                                                                    "current month or full date, including year, "
                                                                    "month and day (please, mind the separators, "
                                                                    "such as ',', '.', ' ', '/', '\\', '-'). You can "
                                                                    "also specify the order of "
                                                                    "information by passing column name after "
                                                                    "the month (e.g. 2021-12-12 amount).\n\n"
                                                                    "COLUMNS AVAILABLE: amount, type, user_time")

    return GET_DAY_LOOKUP


def bot_month_lookup(update: Update, context: CallbackContext):
    context.bot.send_message(chat_id=update.effective_chat.id, text="Please, enter the month you want to get"
                                                                    "information on. You can specify the order of "
                                                                    "information by passing column name after "
                                                                    "the month (e.g. 12 amount).\n\n"
                                                                    "COLUMNS AVAILABLE: amount, type, day, user_time")

    return GET_MONTH_LOOKUP


def bot_check_budget_and_left(update: Update, context: CallbackContext):
    reset_budget(update.effective_user.username.lower())

    planned, spent = get_budget_and_spent(update.effective_user.username.lower())

    context.bot.send_message(chat_id=update.effective_chat.id,
                             text="Your budget is {0} and you've spent {1} (delta is {2}).".format(planned, spent,
                                                                                                   planned - spent))


def bot_categorylookup_receive_category(update: Update, context: CallbackContext):
    user_data = update.message.text

    if bot_check_stop_in_lookup(user_data, update, context):
        return ConversationHandler.END

    user_data = user_data.split()
    order = "amount"
    try:
        order = user_data[1]
    except IndexError:
        pass

    if user_data[0] == "bot_cats":
        data = database.user_help_categories(update.message.from_user.username.lower()).fetched_info
        response = ''
        for row in data:
            response += "{0}\n".format(row[0])
        context.bot.send_message(chat_id=update.effective_chat.id,
                                 text="Here are all the categories that you've logged:\n" + response)
        return GET_CATEGORY

    data = database.lookup(update.message.from_user.username.lower(), "type",
                           user_data[0].lower(), order, to_ret="day").fetched_info

    if not data:
        context.bot.send_message(chat_id=update.effective_chat.id,
                                 text="Hm... I can't find that category, please, try again! If you need help "
                                      "remembering your categories, type 'bot_cats' (though we have no cats, sorry :( "
                                      ").")
        return GET_CATEGORY

    else:
        response = ''
        for row in data:
            response += 'Spent {0} at {1}, your comment: {2}\n'.format(row[0],
                                                                       row[2].strftime("%Y.%m.%d"), row[3])
        context.bot.send_message(chat_id=update.effective_chat.id, text="Here's what I found!:\n" + response)
        return ConversationHandler.END


def bot_daylookup_receive_date(update: Update, context: CallbackContext):
    user_data = update.message.text

    if bot_check_stop_in_lookup(user_data, update, context):
        return ConversationHandler.END

    user_data = user_data.split()
    order = "amount"
    try:
        order = user_data[1]
    except IndexError:
        pass

    data = database.lookup(update.message.from_user.username.lower(), "day", parse_date(user_data[0]),
                           order).fetched_info

    if not data:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Hm... It seems that there's no data from "
                                                                        "specified day, please, try again!")
        return GET_DAY_LOOKUP

    else:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Records found, preparing to send them to you!")

        amount_size = 12
        category_size = 14
        date_size = 12
        comment_size = 13

        for row in data:  # Yes, it's slow, but pretty :3
            amount_size = max(amount_size, len(row[0]))
            category_size = max(category_size, len(row[1]))
            comment_size = max(comment_size, len(row[3]))

        response = "```\n|{0}|{1}|{2}|{3}|\n".format("Amount".center(amount_size, '-'),
                                                     "Category".center(category_size, '-'),
                                                     "Time".center(date_size, '-'), "Comment".center(comment_size, '-'))

        for row in data:
            response += "|{0}|{1}|{2}|{3}|\n".format(row[0].center(amount_size), row[1].center(category_size),
                                                     row[2].strftime("%I:%M:%S").center(date_size),
                                                     row[3].center(comment_size))

        context.bot.send_message(chat_id=update.effective_chat.id, text=response + "\n```", parse_mode='MarkdownV2')
        return ConversationHandler.END


def bot_monthlookup_receive_month(update: Update, context: CallbackContext):
    user_data = update.message.text

    if bot_check_stop_in_lookup(user_data, update, context):
        return ConversationHandler.END

    user_data = user_data.split()
    order = "day"
    try:
        order = user_data[1]
    except IndexError:
        pass

    data = database.lookup_month(update.message.from_user.username.lower(), "EXTRACT(MONTH FROM day)", user_data[0],
                                 order).fetched_info

    if not data:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Hm... It seems that there's no data from "
                                                                        "specified month, please, try again!")
        return GET_DAY_LOOKUP
    else:
        output_table = openpyxl.Workbook()
        ws = output_table.active

        table_headers = ["Amount", "Type", "Date", "Time", "Comment"]
        for i in range(1, 6):
            ws.cell(1, i).value = table_headers[i - 1]
            ws.cell(1, i).border = thin_border

        for i in range(len(data)):
            for j in range(len(data[i])):
                ws.cell(i + 2, j + 1).value = data[i][j]

        output_table.save(update.effective_user.username.lower() + ".xlsx")
        context.bot.send_message(chat_id=update.effective_chat.id, text="Your monthly table is ready! Sending...")
        context.bot.send_document(chat_id=update.effective_chat.id,
                                  document=open(update.effective_user.username.lower() + ".xlsx", "rb"),
                                  filename=update.effective_chat.username.lower() + '.xlsx')
        os.remove(update.effective_user.username.lower() + ".xlsx")


def bot_start(update: Update, context: CallbackContext):
    reset_budget(update.effective_user.username.lower())
    context.bot.send_message(chat_id=update.effective_chat.id, text="Checking if there's already a table for you.")
    try:
        operation = database.init_new_user(update.effective_user.username.lower())
        context.bot.send_message(chat_id=update.effective_chat.id, text=operation.message)  # Created/exists
    except database.psycopg2.Error:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Sorry, something went wrong")


def bot_help(update: Update, context: CallbackContext):
    context.bot.send_message(chat_id=update.effective_chat.id, text="Allowed commands:\n"
                                                                    "/start - use this command if you are a new user\n"
                                                                    "/set_budget [AMOUNT] - set your monthly budget and"
                                                                    " the bot will track amount of money you can spend"
                                                                    " (note: this is inline command)\n"
                                                                    "/check_budget - check your current budget\n"
                                                                    "/add - initiate process of adding new "
                                                                    "transaction\n"
                                                                    "/add_inline - initiate process of adding new"
                                                                    "transaction in a single line\n"
                                                                    "/addhelp - show tips for using 'add' command\n"
                                                                    "/daylookup - get info on particular day\n"
                                                                    "/categorylookup - get info on particular "
                                                                    "category\n"
                                                                    "/monthlookup - get all the info on particular "
                                                                    "month (in format of .xlsx file)\n"
                                                                    "/help - well, don't you know what's that for???\n"
                                                                    "also, if you need to stop lookup'ing or add'ing, "
                                                                    "you can type 'urfin_stop'.")


def bot_addhelp(update: Update, context: CallbackContext):
    context.bot.send_message(chat_id=update.effective_chat.id,
                             text="Okay, here's what's you need to know before use '/add' command:\n"
                                  "If you are asked of transaction cost, please, send it in format of single line of "
                                  "numbers\n"
                                  "If you need to send me type of transaction, "
                                  "feel free to type anything you want\n"
                                  "When the time is asked, I will give you advice on that\n"
                                  "Commentary is also can be anything you want!\n\n"
                                  "Furthermore, you can add your transaction in single-line ('/add_inline'), "
                                  "but the format is very strict:\n\n"
                                  "CONSIDER '__' AS TWO WHITESPACES, IT IS IMPORTANT\n\n"
                                  "/add_inline__COST(line of numbers)__TYPE(line of words)__"
                                  "TIME(hh:mm)__COMMENT(line of words, may be empty)")


def bot_add_inline(update: Update, context: CallbackContext):
    reset_budget(update.effective_user.username.lower())

    msg = update.message.text.split('  ')
    try:
        amount = msg[1]
        transaction_type = msg[2]

        user_time = msg[3]
        now = datetime.datetime.now()
        user_time = datetime.datetime(now.year, now.month, day=now.day, hour=int(user_time[:2]),
                                      minute=int(user_time[3:])).strftime('%Y-%m-%d %H:%M:%S')
        if len(msg) > 4:
            comment = msg[4]
        else:
            comment = ""
        day = str(now.year) + str(now.month) + str(now.day)

        context.bot.send_message(chat_id=update.effective_chat.id, text="Trying to add your record into database.")
        try:
            database.add(update.effective_user.username.lower(), amount, transaction_type, day, user_time, comment)

            database.update_spent(update.effective_user.username.lower(), "EXTRACT(MONTH FROM day)", now.month)
            planned, spent = get_budget_and_spent(update.effective_user.username.lower())
            left_amount = planned - spent if spent < planned else 0

            context.bot.send_message(chat_id=update.effective_chat.id,
                                     text="Done! You have {0} more to spend this month!".format(left_amount))
        except database.psycopg2.Error:
            context.bot.send_message(chat_id=update.effective_chat.id, text="Sorry, something went wrong. :(")
    except IndexError:
        context.bot.send_message(chat_id=update.effective_chat.id, text="It seems, you've entered something in wrong "
                                                                        "format, please, try again.")


def bot_add(update: Update, context: CallbackContext):
    reset_budget(update.effective_user.username.lower())

    context.bot.send_message(chat_id=update.effective_chat.id, text="Arrrr, you want to log some wasted money? "
                                                                    "Please tell me how much.")

    return TYPING_AMOUNT


def bot_add_receive_amount(update: Update, context: CallbackContext):
    amount = update.message.text

    if bot_check_stop_in_lookup(amount, update, context):
        return ConversationHandler.END

    context.bot.send_message(chat_id=update.effective_chat.id, text="Wonderful! Going next: what was the type of the "
                                                                    "spending?")
    context.user_data["amount"] = amount

    return TYPING_TYPE


def bot_add_receive_type(update: Update, context: CallbackContext):
    transaction_type = update.message.text.lower()

    if bot_check_stop_in_lookup(transaction_type, update, context):
        return ConversationHandler.END

    context.user_data["type"] = transaction_type
    context.bot.send_message(chat_id=update.effective_chat.id, text="Now let's talk about time. If you remember "
                                                                    "approximate time of your spending, enter it "
                                                                    "in "
                                                                    "format of 'hh:mm'.")

    return TYPING_TIME


def bot_add_receive_time(update: Update, context: CallbackContext):
    user_time = update.message.text

    if bot_check_stop_in_lookup(user_time, update, context):
        return ConversationHandler.END

    context.user_data["time"] = user_time
    context.bot.send_message(chat_id=update.effective_chat.id, text="Any commentaries? Leave blank if no.")

    return TYPING_COMMENT


def bot_add_receive_comment(update: Update, context: CallbackContext):
    comment = update.message.text

    if bot_check_stop_in_lookup(comment, update, context):
        return ConversationHandler.END

    context.user_data["comment"] = comment
    context.bot.send_message(chat_id=update.effective_chat.id, text="Okay, everything is set up, now I'm trying to "
                                                                    "add your record into database.")

    bot_add_insert(update, context)

    return ConversationHandler.END


def bot_add_insert(update: Update, context: CallbackContext):
    user = context.user_data
    now = datetime.datetime.now()
    day = str(now.year) + str(now.month) + str(now.day)
    user["time"] = datetime.datetime(now.year, now.month, day=now.day, hour=int(user["time"][:2]),
                                     minute=int(user["time"][3:])).strftime('%Y-%m-%d %H:%M:%S')
    try:
        database.add(update.effective_user.username.lower(), user["amount"], user["type"], day, user["time"],
                     user["comment"])

        database.update_spent(update.effective_user.username.lower(), "EXTRACT(MONTH FROM day)", now.month)
        planned, spent = get_budget_and_spent(update.effective_user.username.lower())
        left_amount = planned - spent if spent < planned else 0

        context.bot.send_message(chat_id=update.effective_chat.id,
                                 text="Done! You have {0} more to spend this month!".format(left_amount))
    except database.psycopg2.Error:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Sorry, something went wrong. :(")


def bot_message(update: Update, context: CallbackContext):
    context.bot.send_message(chat_id=update.effective_chat.id, text="Hi! Sorry, but I'm a bit stupid and I understand "
                                                                    "only certain commands, that you can check by "
                                                                    "asking me for help by '/help'.")


def bot_initialize_and_start():
    with open('config/token.txt') as f:
        token = f.readline().strip()
    updater = Updater(token=token)

    dispatcher = updater.dispatcher
    dispatcher.add_handler(CommandHandler('start', bot_start))
    dispatcher.add_handler(CommandHandler('help', bot_help))
    dispatcher.add_handler(CommandHandler('addhelp', bot_addhelp))
    dispatcher.add_handler(CommandHandler('add_inline', bot_add_inline))
    dispatcher.add_handler(CommandHandler('set_budget', bot_set_budget))
    dispatcher.add_handler(CommandHandler('check_budget', bot_check_budget_and_left))

    add_conv = ConversationHandler(entry_points=[CommandHandler('add', bot_add)],
                                   states={
                                       TYPING_AMOUNT: [
                                           MessageHandler(Filters.all,
                                                          bot_add_receive_amount)],
                                       TYPING_TYPE: [
                                           MessageHandler(Filters.text & (~Filters.command), bot_add_receive_type)],
                                       TYPING_TIME: [
                                           MessageHandler(
                                               Filters.text & (~Filters.command) & Filters.regex('\d\d:\d\d'),
                                               bot_add_receive_time)],
                                       TYPING_COMMENT: [
                                           MessageHandler(Filters.text & (~Filters.command), bot_add_receive_comment)]
                                   },
                                   fallbacks=[MessageHandler(~Filters.command, bot_message)])

    day_lookup_conv = ConversationHandler(entry_points=[CommandHandler('daylookup', bot_day_lookup)],
                                          states={
                                              GET_DAY_LOOKUP: [
                                                  MessageHandler(Filters.text & (~Filters.command),
                                                                 bot_daylookup_receive_date)]},
                                          fallbacks=[MessageHandler(~Filters.command, bot_message)])
    category_lookup_conv = ConversationHandler(entry_points=[CommandHandler('categorylookup', bot_category_lookup)],
                                               states={
                                                   GET_CATEGORY: [
                                                       MessageHandler(Filters.text & (~Filters.command),
                                                                      bot_categorylookup_receive_category)]},
                                               fallbacks=[MessageHandler(~Filters.command, bot_message)])
    month_lookup_conv = ConversationHandler(entry_points=[CommandHandler('monthlookup', bot_month_lookup)],
                                            states={
                                                GET_MONTH_LOOKUP: [
                                                    MessageHandler(Filters.text & (~Filters.command),
                                                                   bot_monthlookup_receive_month)]},
                                            fallbacks=[MessageHandler(~Filters.command, bot_message)])

    dispatcher.add_handler(add_conv)
    dispatcher.add_handler(day_lookup_conv)
    dispatcher.add_handler(category_lookup_conv)
    dispatcher.add_handler(month_lookup_conv)

    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    database.init()
    bot_initialize_and_start()
